from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, CreateView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import Project, Company, SystemShare, Activity
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponseRedirect, JsonResponse, FileResponse
from .forms import ProjectForm, CompanyForm
import json
from users.models import User
from django.db import models
from datetime import datetime, timedelta
from django.template.loader import render_to_string

# Constants
SYSTEM_TYPE_CHOICES = {
    '1': 'Database',
    '2': 'VPN',
    '3': 'SERVER',
    '4': 'Application'
}

# Utility functions
def get_system_type_label(system_type):
    return SYSTEM_TYPE_CHOICES.get(system_type, system_type)

def create_success_response(data):
    return JsonResponse({'success': True, **data})

def create_error_response(error_message):
    return JsonResponse({'success': False, 'error': error_message})


class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('keychain:home')
        return render(request, 'keychain/login.html')
    
    def post(self, request):
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f"Hoş geldiniz, {user.username}!")
            print(f"User logged in: {user.username}")  # Debug
            return redirect('keychain:home')
        else:
            messages.error(request, "Kullanıcı adı veya şifre hatalı!")
            return render(request, 'keychain/login.html')


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = 'keychain/home.html'

    def get(self, request, *args, **kwargs):
        print(f"Home view accessed. User authenticated: {request.user.is_authenticated}")  # Debug
        
        # Herkese açık projeler + kendi özel projeleri
        all_projects = Project.objects.filter(
            models.Q(is_private=False) | models.Q(owner=request.user)
        )
        
        user_projects_count = all_projects.count()
        recent_projects = all_projects.order_by('-created_date')[:5]
        
        context = {
            'user': request.user,
            'projects_count': user_projects_count,
            'recent_projects': recent_projects,
        }
        print(f"Rendering home.html for user: {request.user.username}")  # Debug
        return self.render_to_response(context)


class LogoutView(LoginRequiredMixin, View):
    def get(self, request):
        logout(request)
        messages.info(request, "Başarıyla çıkış yaptınız.")
        return redirect('keychain:login')


class ProjectListView(LoginRequiredMixin, ListView):
    model = Project
    template_name = "keychain/project_list.html"
    context_object_name = "projects"

    def get_queryset(self):
        # Herkese açık projeler + kendi özel projeleri
        all_projects = Project.objects.filter(
            models.Q(is_private=False) | models.Q(owner=self.request.user)
        )
        
        # Sıralama parametresini al
        sort_by = self.request.GET.get('sort', 'id')
        order = self.request.GET.get('order', 'asc')
        
        # Sıralama alanını belirle
        if sort_by == 'code':
            sort_field = 'code'
        elif sort_by == 'name':
            sort_field = 'name'
        elif sort_by == 'created_date':
            sort_field = 'created_date'
        elif sort_by == 'modified_date':
            sort_field = 'modified_date'
        else:
            sort_field = 'id'
        
        # Sıralama yönünü belirle
        if order == 'desc':
            sort_field = '-' + sort_field
        
        return all_projects.order_by(sort_field)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = ProjectForm()
        context["current_sort"] = self.request.GET.get('sort', 'id')
        context["current_order"] = self.request.GET.get('order', 'asc')
        return context


class ProjectDetailView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = "keychain/project_detail.html"
    context_object_name = "project"

    def get_queryset(self):
        # Herkese açık projeler + kendi özel projeleri
        return Project.objects.filter(
            models.Q(is_private=False) | models.Q(owner=self.request.user)
        )

class ProjectCreateView(LoginRequiredMixin, CreateView):
    model = Project
    fields = ["code","name", "description"] 
    template_name = "keychain/project_form.html"
    success_url = reverse_lazy("keychain:project_list")

    def form_valid(self, form):
        form.instance.owner = self.request.user  
        messages.success(self.request, "Proje başarıyla oluşturuldu.")
        return super().form_valid(form)


class SistemListView(LoginRequiredMixin, ListView):
    model = Company
    template_name = "keychain/sistem_list.html"
    context_object_name = "systems"

    def get_queryset(self):
        # Herkese açık sistemler + kendi özel sistemleri
        return Company.objects.filter(
            models.Q(is_private=False) | models.Q(owner=self.request.user)
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = CompanyForm()
        
        # Herkese açık projeler + kendi özel projeleri
        context["projects"] = Project.objects.filter(
            models.Q(is_private=False) | models.Q(owner=self.request.user)
        )
        return context


class ProjectCreateAPIView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            project = Project.objects.create(
                code=data.get('code'),
                name=data.get('name'),
                description=data.get('description', ''),
                system_type=data.get('system_type', ''),
                owner=request.user
            )
            
            project.users.add(request.user)
            
            return create_success_response({
                'id': project.id,
                'code': project.code,
                'name': project.name,
                'description': project.description,
                'system_type': get_system_type_label(project.system_type),
                'is_private': project.is_private,
                'status': "Bana Özel" if project.is_private else "Herkese Açık"
            })
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class ProjectUpdateAPIView(LoginRequiredMixin, View):
    def put(self, request, project_id):
        try:
            project = Project.objects.get(id=project_id, owner=request.user)
            data = json.loads(request.body)
            
            project.code = data.get('code', project.code)
            project.name = data.get('name', project.name)
            project.description = data.get('description', project.description)
            project.system_type = data.get('system_type', project.system_type)
            project.save()
            
            return create_success_response({
                'id': project.id,
                'code': project.code,
                'name': project.name,
                'description': project.description,
                'system_type': get_system_type_label(project.system_type),
                'is_private': project.is_private,
                'status': "Bana Özel" if project.is_private else "Herkese Açık"
            })
        except Project.DoesNotExist:
            return create_error_response('Project not found')
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class ProjectDeleteAPIView(LoginRequiredMixin, View):
    def delete(self, request, project_id):
        try:
            project = Project.objects.get(id=project_id, owner=request.user)
            project.delete()
            return JsonResponse({'success': True})
        except Project.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Project not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ProjectTogglePrivacyAPIView(LoginRequiredMixin, View):
    def put(self, request, project_id):
        try:
            project = Project.objects.get(id=project_id, owner=request.user)
            # is_private değerini tersine çevir
            project.is_private = not project.is_private
            project.save()
            
            status = "Bana Özel" if project.is_private else "Herkese Açık"
            return create_success_response({
                'id': project.id,
                'is_private': project.is_private,
                'status': status,
                'message': f'Proje durumu "{status}" olarak güncellendi.'
            })
        except Project.DoesNotExist:
            return create_error_response('Project not found')
        except Exception as e:
            return create_error_response(str(e))


class CompanyCreateAPIView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            project_id = data.get('project_id')
            project = None
            if project_id:
                try:
                    # Herkese açık projeler + kendi özel projeleri
                    project = Project.objects.filter(
                        models.Q(is_private=False) | models.Q(owner=request.user)
                    ).get(id=project_id)
                except Project.DoesNotExist:
                    return create_error_response('Project not found')
            
            company = Company.objects.create(
                project=project,
                name=data.get('name'),
                number=data.get('number'),
                system_type=data.get('system_type'),
                additional_info=data.get('additional_info', {}),
                owner=request.user
            )
            
            company.users.add(request.user)
            
            return create_success_response({
                'id': company.id,
                'project_name': project.name if project else '',
                'name': company.name,
                'number': company.number,
                'system_type': get_system_type_label(company.system_type),
                'is_private': company.is_private,
                'status': "Bana Özel" if company.is_private else "Herkese Açık"
            })
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class CompanyUpdateAPIView(LoginRequiredMixin, View):
    def put(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
            data = json.loads(request.body)
            
            project_id = data.get('project_id')
            if project_id:
                try:
                    # Herkese açık projeler + kendi özel projeleri
                    project = Project.objects.filter(
                        models.Q(is_private=False) | models.Q(owner=request.user)
                    ).get(id=project_id)
                    company.project = project
                except Project.DoesNotExist:
                    return create_error_response('Project not found')
            else:
                company.project = None
            
            company.name = data.get('name', company.name)
            company.number = data.get('number', company.number)
            company.system_type = data.get('system_type', company.system_type)
            company.additional_info = data.get('additional_info', company.additional_info)
            company.save()
            
            return create_success_response({
                'id': company.id,
                'project_name': company.project.name if company.project else '',
                'name': company.name,
                'number': company.number,
                'system_type': get_system_type_label(company.system_type),
                'is_private': company.is_private,
                'status': "Bana Özel" if company.is_private else "Herkese Açık"
            })
        except Company.DoesNotExist:
            return create_error_response('Company not found')
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class CompanyDeleteAPIView(LoginRequiredMixin, View):
    def delete(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id)
            
            # Önce sistemin paylaşımlarını sil
            SystemShare.objects.filter(system=company).delete()
            
            # Sonra sistemi sil
            company.delete()
            return create_success_response({'message': 'Company deleted successfully'})
        except Company.DoesNotExist:
            return create_error_response('Company not found')
        except Exception as e:
            return create_error_response(str(e))


class CompanyTogglePrivacyAPIView(LoginRequiredMixin, View):
    def put(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id, owner=request.user)
            # is_private değerini tersine çevir
            company.is_private = not company.is_private
            company.save()
            
            status = "Bana Özel" if company.is_private else "Herkese Açık"
            return create_success_response({
                'id': company.id,
                'is_private': company.is_private,
                'status': status,
                'message': f'Sistem durumu "{status}" olarak güncellendi.'
            })
        except Company.DoesNotExist:
            return create_error_response('System not found')
        except Exception as e:
            return create_error_response(str(e))


class CompanyDetailAPIView(LoginRequiredMixin, View):
    def get(self, request, company_id):
        try:
            # Herkese açık sistemler + kendi özel sistemleri
            company = Company.objects.filter(
                models.Q(is_private=False) | models.Q(owner=request.user)
            ).get(id=company_id)

            # Sistem tipine göre ek bilgileri hazırla
            additional_info = company.additional_info or {}
            system_type = get_system_type_label(company.system_type)
            
            # Sistem tipine göre detaylı bilgileri ekle
            detailed_info = {
                'id': company.id,
                'project_name': company.project.name if company.project else '',
                'name': company.name,
                'number': company.number,
                'system_type': system_type,
                'additional_info': additional_info
            }

            # Sistem tipine göre özel alanları ekle
            if system_type == 'Database':
                detailed_info.update({
                    'host': additional_info.get('host', ''),
                    'port': additional_info.get('port', ''),
                    'username': additional_info.get('username', ''),
                    'password': additional_info.get('password', '')
                })
            elif system_type == 'VPN':
                detailed_info.update({
                    'server': additional_info.get('server', ''),
                    'port': additional_info.get('port', ''),
                    'username': additional_info.get('username', ''),
                    'password': additional_info.get('password', '')
                })
            elif system_type == 'SERVER':
                detailed_info.update({
                    'ip': additional_info.get('ip', ''),
                    'os': additional_info.get('os', ''),
                    'username': additional_info.get('username', ''),
                    'password': additional_info.get('password', '')
                })
            elif system_type == 'Application':
                detailed_info.update({
                    'url': additional_info.get('url', ''),
                    'version': additional_info.get('version', ''),
                    'username': additional_info.get('username', ''),
                    'password': additional_info.get('password', '')
                })

            return create_success_response(detailed_info)
        except Company.DoesNotExist:
            return create_error_response('System not found')
        except Exception as e:
            return create_error_response(str(e))


@method_decorator(csrf_exempt, name='dispatch')
class SystemShareAPIView(LoginRequiredMixin, View):
    def get(self, request):
        try:
            # Tüm kullanıcıları getir (mevcut kullanıcı hariç)
            users = User.objects.exclude(id=request.user.id)
            users_data = [{'id': user.id, 'username': user.username} for user in users]
            return JsonResponse({'success': True, 'users': users_data})
        except Exception as e:
            return create_error_response(str(e))

    def post(self, request):
        # Sistem paylaşım yetkisi kontrolü
        if not request.user.can_share_systems:
            return create_error_response('Sistem paylaşım yetkiniz bulunmamaktadır.')
        
        try:
            data = json.loads(request.body)
            system_id = data.get('system_id')
            is_public = data.get('is_public', False)
            shared_with_ids = data.get('shared_with', [])

            system = request.user.companies.get(id=system_id)
            
            # Mevcut paylaşımı kontrol et ve güncelle
            share, created = SystemShare.objects.get_or_create(
                system=system,
                shared_by=request.user,
                defaults={'is_public': is_public}
            )
            
            if not created:
                share.is_public = is_public
                share.save()
            
            # Paylaşım listesini güncelle
            if not is_public:
                share.shared_with.clear()
                for user_id in shared_with_ids:
                    try:
                        user = User.objects.get(id=user_id)
                        share.shared_with.add(user)
                    except User.DoesNotExist:
                        continue
            else:
                share.shared_with.clear()

            return create_success_response({
                'message': 'System shared successfully',
                'is_public': share.is_public,
                'shared_with': [user.id for user in share.shared_with.all()]
            })
        except Company.DoesNotExist:
            return create_error_response('System not found')
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class ActivityListView(LoginRequiredMixin, ListView):
    model = Activity
    template_name = "keychain/activity_list.html"
    context_object_name = "activities"

    def get_queryset(self):
        # Kullanıcının sahip olduğu veya atanmış olduğu faaliyetler
        user_activities = Activity.objects.filter(
            models.Q(owner=self.request.user) |
            models.Q(primary_person=self.request.user) |
            models.Q(secondary_person=self.request.user)
        ).distinct()
        
        # Arama filtresi
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            user_activities = user_activities.filter(
                models.Q(activity_name__icontains=search_query) |
                models.Q(project__code__icontains=search_query) |
                models.Q(project__name__icontains=search_query) |
                models.Q(primary_person__username__icontains=search_query) |
                models.Q(secondary_person__username__icontains=search_query)
            )
        
        # Tarih filtrelerini uygula
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                user_activities = user_activities.filter(activity_date__gte=start_date)
            except ValueError:
                pass  # Geçersiz tarih formatı, filtreyi uygulama
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                user_activities = user_activities.filter(activity_date__lte=end_date)
            except ValueError:
                pass  # Geçersiz tarih formatı, filtreyi uygulama
        
        # Sıralama parametresini al
        sort_by = self.request.GET.get('sort', 'activity_date')
        order = self.request.GET.get('order', 'desc')
        
        # Sıralama alanını belirle
        if sort_by == 'project_code':
            sort_field = 'project__code'
        elif sort_by == 'project_name':
            sort_field = 'project__name'
        elif sort_by == 'activity_name':
            sort_field = 'activity_name'
        elif sort_by == 'duration':
            sort_field = 'duration'
        elif sort_by == 'is_billable':
            sort_field = 'is_billable'
        elif sort_by == 'primary_person':
            sort_field = 'primary_person__username'
        elif sort_by == 'secondary_person':
            sort_field = 'secondary_person__username'
        elif sort_by == 'activity_date':
            sort_field = 'activity_date'
        else:
            sort_field = 'activity_date'
        
        # Sıralama yönünü belirle
        if order == 'asc':
            sort_field = sort_field
        else:
            sort_field = '-' + sort_field
        
        return user_activities.select_related('project', 'primary_person', 'secondary_person').order_by(sort_field)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Kullanıcının erişebildiği projeler
        context["projects"] = Project.objects.filter(
            models.Q(is_private=False) | models.Q(owner=self.request.user)
        )
        # Tüm kullanıcılar (faaliyet ataması için)
        context["users"] = User.objects.all()
        context["current_sort"] = self.request.GET.get('sort', 'activity_date')
        context["current_order"] = self.request.GET.get('order', 'desc')
        # Mevcut tarih filtreleri
        context["start_date"] = self.request.GET.get('start_date', '')
        context["end_date"] = self.request.GET.get('end_date', '')
        # Mevcut arama terimi
        context["current_search"] = self.request.GET.get('search', '')
        return context


class ActivityCreateAPIView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            # Form data mı JSON data mı kontrol et
            if request.content_type and 'application/json' in request.content_type:
                data = json.loads(request.body)
            else:
                # Form data (dosya upload için)
                data = {
                    'project_id': request.POST.get('project_id'),
                    'activity_name': request.POST.get('activity_name'),
                    'duration': request.POST.get('duration'),
                    'is_billable': request.POST.get('is_billable') == 'true',
                    'primary_person_id': request.POST.get('primary_person_id'),
                    'secondary_person_id': request.POST.get('secondary_person_id'),
                    'activity_date': request.POST.get('activity_date')
                }
            
            # Proje kontrolü
            project_id = data.get('project_id')
            try:
                project = Project.objects.filter(
                    models.Q(is_private=False) | models.Q(owner=request.user)
                ).get(id=project_id)
            except Project.DoesNotExist:
                return create_error_response('Project not found')
            
            # Kullanıcı kontrolü
            primary_person_id = data.get('primary_person_id')
            secondary_person_id = data.get('secondary_person_id')
            
            try:
                primary_person = User.objects.get(id=primary_person_id)
            except User.DoesNotExist:
                return create_error_response('Primary person not found')
            
            secondary_person = None
            if secondary_person_id:
                try:
                    secondary_person = User.objects.get(id=secondary_person_id)
                except User.DoesNotExist:
                    return create_error_response('Secondary person not found')
            
            # Süreyi saat cinsinden al ve timedelta'ya çevir
            duration_hours = data.get('duration', 0)
            try:
                duration_hours = float(duration_hours)
                if duration_hours < 0:
                    return create_error_response('Duration must be positive')
                # Saat ve dakikaya çevir
                hours = int(duration_hours)
                minutes = int((duration_hours - hours) * 60)
                duration = timedelta(hours=hours, minutes=minutes)
            except (ValueError, TypeError):
                return create_error_response('Invalid duration format. Must be a number')
            
            # Tarih formatını kontrol et
            activity_date_str = data.get('activity_date')
            try:
                activity_date = datetime.strptime(activity_date_str, '%Y-%m-%d').date()
            except ValueError:
                return create_error_response('Invalid date format. Use YYYY-MM-DD format')
            
            activity = Activity.objects.create(
                project=project,
                activity_name=data.get('activity_name'),
                duration=duration,
                is_billable=data.get('is_billable', True),
                primary_person=primary_person,
                secondary_person=secondary_person,
                activity_date=activity_date,
                owner=request.user
            )
            
            # Dosya varsa kaydet
            if 'attachment' in request.FILES:
                activity.attachment = request.FILES['attachment']
                activity.save()
            
            return create_success_response({
                'id': activity.id,
                'project_code': activity.project.code,
                'project_name': activity.project.name,
                'activity_name': activity.activity_name,
                'duration': str(activity.duration),
                'duration_hours': activity.duration_hours,
                'is_billable': activity.is_billable,
                'billable_status': activity.billable_status,
                'primary_person': activity.primary_person.username,
                'secondary_person': activity.secondary_person.username if activity.secondary_person else '',
                'activity_date': activity.activity_date.strftime('%Y-%m-%d')
            })
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class ActivityUpdateAPIView(LoginRequiredMixin, View):
    def post(self, request, activity_id):
        # POST methodunu PUT olarak işle
        return self.put(request, activity_id)
    
    def put(self, request, activity_id):
        try:
            activity = Activity.objects.get(id=activity_id, owner=request.user)
            
            # Form data mı JSON data mı kontrol et
            if request.content_type and 'application/json' in request.content_type:
                data = json.loads(request.body)
            else:
                # Form data (dosya upload için)
                data = {
                    'project_id': request.POST.get('project_id'),
                    'activity_name': request.POST.get('activity_name'),
                    'duration': request.POST.get('duration'),
                    'is_billable': request.POST.get('is_billable') == 'true',
                    'primary_person_id': request.POST.get('primary_person_id'),
                    'secondary_person_id': request.POST.get('secondary_person_id'),
                    'activity_date': request.POST.get('activity_date')
                }
            
            # Proje kontrolü
            project_id = data.get('project_id')
            if project_id:
                try:
                    project = Project.objects.filter(
                        models.Q(is_private=False) | models.Q(owner=request.user)
                    ).get(id=project_id)
                    activity.project = project
                except Project.DoesNotExist:
                    return create_error_response('Project not found')
            
            # Kullanıcı kontrolü
            primary_person_id = data.get('primary_person_id')
            if primary_person_id:
                try:
                    activity.primary_person = User.objects.get(id=primary_person_id)
                except User.DoesNotExist:
                    return create_error_response('Primary person not found')
            
            secondary_person_id = data.get('secondary_person_id')
            if secondary_person_id:
                try:
                    activity.secondary_person = User.objects.get(id=secondary_person_id)
                except User.DoesNotExist:
                    return create_error_response('Secondary person not found')
            elif secondary_person_id == '':
                activity.secondary_person = None
            
            # Süre formatını kontrol et
            duration_hours = data.get('duration')
            if duration_hours is not None:
                try:
                    duration_hours = float(duration_hours)
                    if duration_hours < 0:
                        return create_error_response('Duration must be positive')
                    # Saat ve dakikaya çevir
                    hours = int(duration_hours)
                    minutes = int((duration_hours - hours) * 60)
                    activity.duration = timedelta(hours=hours, minutes=minutes)
                except (ValueError, TypeError):
                    return create_error_response('Invalid duration format. Must be a number')
            
            # Tarih formatını kontrol et
            activity_date_str = data.get('activity_date')
            if activity_date_str:
                try:
                    activity.activity_date = datetime.strptime(activity_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return create_error_response('Invalid date format. Use YYYY-MM-DD format')
            
            # Diğer alanları güncelle
            activity.activity_name = data.get('activity_name', activity.activity_name)
            activity.is_billable = data.get('is_billable', activity.is_billable)
            
            # Dosya güncelleme
            if 'attachment' in request.FILES:
                # Eski dosyayı sil (varsa)
                if activity.attachment:
                    try:
                        activity.attachment.delete()
                    except:
                        pass
                # Yeni dosyayı kaydet
                activity.attachment = request.FILES['attachment']
            elif request.POST.get('remove_attachment') == 'true':
                # Dosyayı kaldır
                if activity.attachment:
                    try:
                        activity.attachment.delete()
                    except:
                        pass
                    activity.attachment = None
            
            activity.save()
            
            return create_success_response({
                'id': activity.id,
                'project_code': activity.project.code,
                'project_name': activity.project.name,
                'activity_name': activity.activity_name,
                'duration': str(activity.duration),
                'duration_hours': activity.duration_hours,
                'is_billable': activity.is_billable,
                'billable_status': activity.billable_status,
                'primary_person': activity.primary_person.username,
                'secondary_person': activity.secondary_person.username if activity.secondary_person else '',
                'activity_date': activity.activity_date.strftime('%Y-%m-%d')
            })
        except Activity.DoesNotExist:
            return create_error_response('Activity not found')
        except json.JSONDecodeError:
            return create_error_response('Invalid JSON data')
        except Exception as e:
            return create_error_response(str(e))


class ActivityDeleteAPIView(LoginRequiredMixin, View):
    def delete(self, request, activity_id):
        try:
            activity = Activity.objects.get(id=activity_id, owner=request.user)
            activity.delete()
            return JsonResponse({'success': True})
        except Activity.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Activity not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ActivityExportExcelView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
            import io
            
            activity_ids = request.POST.getlist('activity_ids')
            if not activity_ids:
                return JsonResponse({'success': False, 'error': 'Seçili faaliyet bulunamadı'})
            
            # Kullanıcının sahip olduğu veya atanmış olduğu faaliyetler
            activities = Activity.objects.filter(
                id__in=activity_ids
            ).filter(
                models.Q(owner=request.user) | 
                models.Q(primary_person=request.user) | 
                models.Q(secondary_person=request.user)
            ).select_related('project', 'primary_person', 'secondary_person').order_by('activity_date')
            
            if not activities:
                return JsonResponse({'success': False, 'error': 'Seçili faaliyet bulunamadı'})
            
            # Excel workbook oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Faaliyetler"
            
            # Header stilleri
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Header'ları yazd     
            headers = ["Tarih", "Proje Kodu", "Proje Adı", "Faaliyet", "Süre", "Faturlanabilirlik", "Kişi"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Verileri yazd    
            for row, activity in enumerate(activities, 2):
                ws.cell(row=row, column=1, value=activity.activity_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=activity.project.code)
                ws.cell(row=row, column=3, value=activity.project.name)
                ws.cell(row=row, column=4, value=activity.activity_name)
                ws.cell(row=row, column=5, value=f"{activity.duration_hours:.1f} saat")
                ws.cell(row=row, column=6, value="Evet" if activity.is_billable else "Hayır")
                ws.cell(row=row, column=7, value=activity.primary_person.username)
            
            # Sütun genişliklerini ayarla
            column_widths = [12, 15, 25, 30, 10, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            
            # Response oluştur
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="faaliyetler.xlsx"'
            
            # Excel dosyasını response'a kaydet
            with io.BytesIO() as buffer:
                wb.save(buffer)
                response.write(buffer.getvalue())
            
            return response
            
        except ImportError:
            return JsonResponse({'success': False, 'error': 'openpyxl kütüphanesi yüklü değil'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ActivityExportPDFView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            import pdfkit
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            from django.conf import settings
            import os
            
            activity_ids = request.POST.getlist('activity_ids')
            if not activity_ids:
                return JsonResponse({'success': False, 'error': 'Seçili faaliyet bulunamadı'})
            
            # Kullanıcının sahip olduğu veya atanmış olduğu faaliyetler
            activities = Activity.objects.filter(
                id__in=activity_ids
            ).filter(
                models.Q(owner=request.user) | 
                models.Q(primary_person=request.user) | 
                models.Q(secondary_person=request.user)
            ).select_related('project', 'primary_person', 'secondary_person').order_by('activity_date')
            
            # Queryset'i listeye çevir ve kontrol et
            activities_list = list(activities)
            if not activities_list:
                return JsonResponse({'success': False, 'error': 'Seçili faaliyet bulunamadı'})
            
            # Toplam süre hesaplaması
            total_hours = sum(activity.duration_hours for activity in activities_list)
            billable_hours = sum(activity.duration_hours for activity in activities_list if activity.is_billable)
            non_billable_hours = total_hours - billable_hours
            billing_ratio = (billable_hours / total_hours * 100) if total_hours > 0 else 0
            
            # Tarih aralığı
            first_date = activities_list[0].activity_date
            last_date = activities_list[-1].activity_date
            
            # Logo'yu base64 olarak hazırla
            logo_base64 = None
            try:
                import base64
                logo_path = os.path.join(settings.BASE_DIR, 'keychain', 'static', 'keychain', 'images', 'logo.png')
                if os.path.exists(logo_path):
                    with open(logo_path, 'rb') as logo_file:
                        logo_data = base64.b64encode(logo_file.read()).decode('utf-8')
                        logo_base64 = f"data:image/png;base64,{logo_data}"
            except Exception as e:
                pass
            
            # Template context'i hazırla
            context = {
                'activities': activities_list,
                'total_activities': len(activities_list),
                'total_hours': total_hours,
                'billable_hours': billable_hours,
                'non_billable_hours': non_billable_hours,
                'billing_ratio': billing_ratio,
                'first_date': first_date,
                'last_date': last_date,
                'generated_date': datetime.now(),
                'user': request.user,
                'logo_base64': logo_base64,
            }
            
            # HTML template'ini render et
            html_content = render_to_string('keychain/pdf_template.html', context)
            
            # wkhtmltopdf ayarları
            options = {
                'page-size': 'A4',
                'margin-top': '2cm',
                'margin-right': '2cm',
                'margin-bottom': '2cm',
                'margin-left': '2cm',
                'encoding': "UTF-8",
                'custom-header': [
                    ('Accept-Encoding', 'gzip')
                ],
                'no-outline': None,
                'enable-local-file-access': None
            }
            
            # wkhtmltopdf path'ini bul
            wkhtmltopdf_path = None
            possible_paths = [
                r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
                r'C:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe',
                'wkhtmltopdf'  # PATH'de varsa
            ]
            
            for path in possible_paths:
                if os.path.exists(path) or path == 'wkhtmltopdf':
                    wkhtmltopdf_path = path
                    break
            
            if not wkhtmltopdf_path:
                return JsonResponse({
                    'success': False, 
                    'error': 'wkhtmltopdf bulunamadı. Lütfen https://wkhtmltopdf.org/downloads.html adresinden indirip kurun.'
                })
            
            # PDF konfigürasyonu
            config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
            
            try:
                # HTML'i PDF'e çevir
                pdf_content = pdfkit.from_string(html_content, False, options=options, configuration=config)
                
                # Response oluştur
                response = HttpResponse(pdf_content, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="faaliyet_raporu_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
                return response
                
            except Exception as pdf_error:
                return JsonResponse({
                    'success': False, 
                    'error': f'PDF oluşturma hatası: {str(pdf_error)}'
                })
            
        except ImportError:
            return JsonResponse({'success': False, 'error': 'pdfkit kütüphanesi yüklü değil'})
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False, 
                'error': f'Beklenmeyen hata: {str(e)}',
                'traceback': traceback.format_exc()
            })


class ActivityExportPDFTestView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.colors import HexColor
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER
            from django.http import HttpResponse
            import io
            import os
            from django.conf import settings
            
            # PDF buffer oluştur
            pdf_buffer = io.BytesIO()
            
            # PDF dokument oluştur
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Stiller
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=HexColor('#f2b100'),
                alignment=TA_CENTER,
                spaceAfter=20,
                fontName='Helvetica-Bold'
            )
            
            # PDF içeriği
            story = []
            
            # Logo ekle (varsa)
            logo_path = os.path.join(settings.BASE_DIR, 'keychain', 'static', 'keychain', 'images', 'qvaultlogo.png')
            if os.path.exists(logo_path):
                try:
                    logo = Image(logo_path, width=100, height=50)
                    logo.hAlign = 'CENTER'
                    story.append(logo)
                    story.append(Spacer(1, 10))
                except:
                    pass
            
            # Test başlığı
            story.append(Paragraph("TEST PDF - ReportLab ile Oluşturuldu", title_style))
            story.append(Paragraph(f"Oluşturulma Zamanı: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            
            # PDF oluştur
            doc.build(story)
            
            # Response oluştur
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="test.pdf"'
            return response
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ActivityAttachmentDownloadView(LoginRequiredMixin, View):
    """Faaliyet dosyalarını indirme görünümü"""
    def get(self, request, activity_id):
        try:
            # Kullanıcının erişebildiği faaliyetleri kontrol et
            activity = Activity.objects.filter(
                models.Q(owner=request.user) | 
                models.Q(primary_person=request.user) | 
                models.Q(secondary_person=request.user)
            ).get(id=activity_id)
            
            if not activity.attachment:
                return JsonResponse({'success': False, 'error': 'Bu faaliyete ait dosya bulunamadı'})
            
            from django.http import FileResponse
            import os
            
            # Dosya yolunu al
            file_path = activity.attachment.path
            
            # Dosya var mı kontrol et
            if not os.path.exists(file_path):
                return JsonResponse({'success': False, 'error': 'Dosya sistemde bulunamadı'})
            
            # Dosya adını al (sadece dosya adı, yol olmadan)
            filename = os.path.basename(activity.attachment.name)
            
            # Dosyayı indir
            response = FileResponse(
                open(file_path, 'rb'),
                as_attachment=True,
                filename=filename
            )
            
            return response
            
        except Activity.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Faaliyet bulunamadı veya erişim yetkiniz yok'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})